import io
import sqlite3

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from db.init_db import get_connection
from db import crud


def _fetch_session_bundle(session_id: int) -> dict:
    with get_connection() as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute(
            """
            SELECT
                s.id,
                s.timestamp AS datetime,
                p.name AS patient_name,
                e.name AS exercise_name,
                s.notes
            FROM sessions s
            LEFT JOIN patients  p ON p.id = s.patient_id
            LEFT JOIN exercises e ON e.id = s.exercise_id
            WHERE s.id = ?
            """,
            (session_id,),
        )
        row = cur.fetchone()
        if not row:
            raise ValueError(f"Session {session_id} not found")
        return dict(row)


# Column order: metadata first, then positions, then angles
_PREFERRED_COLUMN_ORDER = [
    "Registro (secuencia)",
    "Secuencia",
    "patient_name",
    "datetime",
    "session_id",
    "time_seconds",
    # Right arm
    "shoulder_x_r", "shoulder_y_r",
    "elbow_x_r",    "elbow_y_r",
    "wrist_x_r",    "wrist_y_r",
    # Left arm
    "shoulder_x_l", "shoulder_y_l",
    "elbow_x_l",    "elbow_y_l",
    "wrist_x_l",    "wrist_y_l",
    # Right leg
    "hip_x_r",      "hip_y_r",
    "knee_x_r",     "knee_y_r",
    "ankle_x_r",    "ankle_y_r",
    # Left leg
    "hip_x_l",      "hip_y_l",
    "knee_x_l",     "knee_y_l",
    "ankle_x_l",    "ankle_y_l",
    # Angles
    "angle_arm_r",
    "angle_arm_l",
    "angle_leg_r",
    "angle_leg_l",
]

_HEADER_FILL   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
_META_FILL     = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
_ANGLE_FILL    = PatternFill("solid", start_color="375623", end_color="375623")
_WHITE_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_DATA_FONT     = Font(name="Arial", size=9)

_ANGLE_COLS    = {"angle_arm_r", "angle_arm_l", "angle_leg_r", "angle_leg_l"}
_META_COLS     = {"Registro (secuencia)", "Secuencia", "patient_name", "datetime", "session_id", "time_seconds"}


def generate_session_excel(session_id: int) -> bytes:
    bundle = _fetch_session_bundle(session_id)
    rows   = crud.get_movement_data_by_session(session_id)

    if not rows:
        raise ValueError(f"No movement data found for session {session_id}")

    df = pd.DataFrame(rows)

    # Inject session metadata into every row
    df.insert(0, "patient_name", bundle["patient_name"] or "")
    df.insert(1, "datetime",     bundle["datetime"]     or "")

    # Registro (secuencia) = row index (0-based), Secuencia = frame number
    df.insert(0, "Secuencia",            df["frame"])
    df.insert(0, "Registro (secuencia)", range(len(df)))

    # Drop original frame and id columns (redundant)
    df.drop(columns=[c for c in ("frame", "id") if c in df.columns], inplace=True)

    # Reorder columns: preferred order first, then any extra columns appended
    present   = set(df.columns)
    ordered   = [c for c in _PREFERRED_COLUMN_ORDER if c in present]
    extras    = [c for c in df.columns if c not in set(ordered)]
    df        = df[ordered + extras]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Sesion_{session_id}"

    # --- Header row ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if col_name in _ANGLE_COLS:
            cell.fill = _ANGLE_FILL
        elif col_name in _META_COLS:
            cell.fill = _META_FILL
        else:
            cell.fill = _HEADER_FILL

    ws.row_dimensions[1].height = 30

    # --- Data rows ---
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = _DATA_FONT
            cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")

            # Subtle zebra striping
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color="F2F7FC", end_color="F2F7FC")

    # --- Column widths ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        if col_name in ("patient_name", "datetime"):
            ws.column_dimensions[letter].width = 22
        elif col_name in ("Registro (secuencia)", "Secuencia", "time_seconds"):
            ws.column_dimensions[letter].width = 12
        elif col_name in _ANGLE_COLS:
            ws.column_dimensions[letter].width = 14
        else:
            ws.column_dimensions[letter].width = 11

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Resumen")
    summary_data = [
        ("Sesión ID",    session_id),
        ("Paciente",     bundle["patient_name"] or "—"),
        ("Ejercicio",    bundle["exercise_name"] or "—"),
        ("Fecha",        bundle["datetime"]      or "—"),
        ("Notas",        bundle["notes"]         or "—"),
        ("Total frames", len(df)),
    ]
    for r, (label, value) in enumerate(summary_data, start=1):
        lbl_cell = ws2.cell(row=r, column=1, value=label)
        lbl_cell.font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=r, column=2, value=value).font = Font(name="Arial", size=10)

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()